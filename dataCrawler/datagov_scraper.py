"""
data.gov 完整爬虫
================
使用美国政府开放数据平台 CKAN API，实现：
  - 数据集搜索（关键字、分类、组织）
  - 数据集详情抓取
  - 资源元数据采集
  - 组织 / 标签统计
  - 结构化输出（JSON / CSV / Excel）
  - 速率限制 + 重试 + 日志

API 文档: https://catalog.data.gov/api/3
"""

import csv
import json
import logging
import os
import sys
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional
from urllib.parse import urlencode

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ─── 日志配置 ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("datagov_scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("datagov")

# ─── 常量 ────────────────────────────────────────────────────────────────────
BASE_URL   = "https://catalog.data.gov/api/3/action"
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)


# ─── 数据模型 ─────────────────────────────────────────────────────────────────
@dataclass
class Resource:
    id: str
    name: str
    format: str
    url: str
    size: Optional[int]
    mimetype: Optional[str]
    created: Optional[str]
    last_modified: Optional[str]
    description: Optional[str]


@dataclass
class Dataset:
    id: str
    name: str                        # slug
    title: str
    notes: str                       # description
    organization: Optional[str]
    org_type: Optional[str]
    license_title: Optional[str]
    license_url: Optional[str]
    tags: List[str]
    groups: List[str]
    resources: List[Resource]
    metadata_created: Optional[str]
    metadata_modified: Optional[str]
    num_resources: int
    num_tags: int
    state: str
    url: str                         # portal page URL
    extras: Dict[str, str]           # custom key-value pairs


@dataclass
class SearchResult:
    query: str
    total_count: int
    fetched_count: int
    fetch_time_sec: float
    datasets: List[Dataset] = field(default_factory=list)


# ─── HTTP 会话（重试 + 超时） ──────────────────────────────────────────────────
def _make_session(retries: int = 5, backoff: float = 1.0) -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=retries,
        backoff_factor=backoff,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({
        "User-Agent": "DataGovScraper/1.0 (educational; +https://github.com/example)",
        "Accept": "application/json",
    })
    return session


# ─── CKAN API 客户端 ──────────────────────────────────────────────────────────
class DataGovClient:
    """
    封装 data.gov CKAN API 的所有调用，处理分页、速率限制与错误。
    """

    def __init__(
        self,
        base_url: str = BASE_URL,
        rate_limit_delay: float = 0.3,   # 每次请求间隔（秒）
        timeout: int = 30,
    ):
        self.base_url = base_url
        self.delay   = rate_limit_delay
        self.timeout = timeout
        self.session = _make_session()
        self._last_request_time: float = 0.0

    # ── 底层 GET ─────────────────────────────────────────────────────────────
    def _get(self, endpoint: str, params: Dict[str, Any] = None) -> Dict:
        """统一的 GET 请求，含速率控制。"""
        elapsed = time.time() - self._last_request_time
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)

        url = f"{self.base_url}/{endpoint}"
        log.debug("GET %s  params=%s", url, params)
        try:
            resp = self.session.get(url, params=params, timeout=self.timeout)
            resp.raise_for_status()
            data = resp.json()
            if not data.get("success"):
                raise ValueError(f"API 返回 success=false: {data.get('error')}")
            return data["result"]
        except requests.exceptions.HTTPError as e:
            log.error("HTTP 错误 %s: %s", url, e)
            raise
        except requests.exceptions.ConnectionError as e:
            log.error("连接错误: %s", e)
            raise
        finally:
            self._last_request_time = time.time()

    # ── 搜索数据集 ───────────────────────────────────────────────────────────
    def search_datasets(
        self,
        query: str = "*:*",
        fq: str = "",
        sort: str = "metadata_modified desc",
        rows: int = 20,
        start: int = 0,
    ) -> Dict:
        """
        调用 package_search API。
        query: Solr 全文检索字符串
        fq:    过滤查询，如 'organization:nasa-gov' 或 'res_format:CSV'
        """
        params = {"q": query, "sort": sort, "rows": rows, "start": start}
        if fq:
            params["fq"] = fq
        return self._get("package_search", params)

    def iter_datasets(
        self,
        query: str = "*:*",
        fq: str = "",
        max_results: int = 100,
        page_size: int = 20,
    ) -> Iterator[Dict]:
        """分页迭代数据集原始字典，最多返回 max_results 条。"""
        fetched = 0
        start   = 0
        while fetched < max_results:
            rows = min(page_size, max_results - fetched)
            result = self.search_datasets(query=query, fq=fq, rows=rows, start=start)
            items  = result.get("results", [])
            if not items:
                break
            for item in items:
                yield item
                fetched += 1
                if fetched >= max_results:
                    break
            start += len(items)
            log.info("已获取 %d / %d 条", fetched, min(max_results, result.get("count", 0)))

    # ── 数据集详情 ───────────────────────────────────────────────────────────
    def get_dataset(self, name_or_id: str) -> Dict:
        """获取单个数据集完整信息。"""
        return self._get("package_show", {"id": name_or_id})

    # ── 组织列表 ─────────────────────────────────────────────────────────────
    def list_organizations(
        self, all_fields: bool = True, limit: int = 200
    ) -> List[Dict]:
        result = self._get(
            "organization_list",
            {"all_fields": str(all_fields).lower(), "limit": limit, "include_dataset_count": "true"},
        )
        return result if isinstance(result, list) else []

    # ── 标签列表 ─────────────────────────────────────────────────────────────
    def list_tags(self, query: str = "", limit: int = 200) -> List[Dict]:
        params: Dict[str, Any] = {"all_fields": "true", "limit": limit}
        if query:
            params["query"] = query
        result = self._get("tag_list", params)
        return result if isinstance(result, list) else []

    # ── 数据格式统计 ─────────────────────────────────────────────────────────
    def facet_formats(self, query: str = "*:*", top_n: int = 30) -> Dict[str, int]:
        """返回最常见的资源格式及其数量。"""
        result = self._get(
            "package_search",
            {"q": query, "rows": 0, "facet": "true",
             "facet.field": '["res_format"]', "facet.limit": top_n},
        )
        facets = result.get("search_facets", {})
        fmt_facet = facets.get("res_format", {}).get("items", [])
        return {item["name"]: item["count"] for item in fmt_facet}

    # ── 组织统计 ─────────────────────────────────────────────────────────────
    def facet_organizations(self, query: str = "*:*", top_n: int = 30) -> Dict[str, int]:
        result = self._get(
            "package_search",
            {"q": query, "rows": 0, "facet": "true",
             "facet.field": '["organization"]', "facet.limit": top_n},
        )
        facets = result.get("search_facets", {})
        org_facet = facets.get("organization", {}).get("items", [])
        return {item["display_name"]: item["count"] for item in org_facet}


# ─── 数据解析（原始字典 → 数据类） ────────────────────────────────────────────
def _parse_resource(r: Dict) -> Resource:
    return Resource(
        id=r.get("id", ""),
        name=r.get("name") or r.get("description", "Unnamed"),
        format=r.get("format", ""),
        url=r.get("url", ""),
        size=r.get("size"),
        mimetype=r.get("mimetype"),
        created=r.get("created"),
        last_modified=r.get("last_modified"),
        description=r.get("description"),
    )


def _parse_dataset(d: Dict) -> Dataset:
    org   = d.get("organization") or {}
    extras = {e["key"]: e["value"] for e in d.get("extras", []) if "key" in e}
    return Dataset(
        id=d.get("id", ""),
        name=d.get("name", ""),
        title=d.get("title", ""),
        notes=(d.get("notes") or "").strip(),
        organization=org.get("title") or org.get("name"),
        org_type=org.get("type"),
        license_title=d.get("license_title"),
        license_url=d.get("license_url"),
        tags=[t["name"] for t in d.get("tags", [])],
        groups=[g.get("display_name") or g.get("name", "") for g in d.get("groups", [])],
        resources=[_parse_resource(r) for r in d.get("resources", [])],
        metadata_created=d.get("metadata_created"),
        metadata_modified=d.get("metadata_modified"),
        num_resources=d.get("num_resources", 0),
        num_tags=len(d.get("tags", [])),
        state=d.get("state", ""),
        url=f"https://catalog.data.gov/dataset/{d.get('name', '')}",
        extras=extras,
    )


# ─── 爬虫主控器 ───────────────────────────────────────────────────────────────
class DataGovScraper:
    """
    高层封装：搜索 → 解析 → 输出。
    """

    def __init__(self, rate_limit_delay: float = 0.5):
        self.client = DataGovClient(rate_limit_delay=rate_limit_delay)

    # ── 搜索并返回结构化结果 ─────────────────────────────────────────────────
    def search(
        self,
        query: str = "*:*",
        fq: str = "",
        max_results: int = 50,
    ) -> SearchResult:
        log.info("开始搜索: query=%r  fq=%r  max=%d", query, fq, max_results)
        start_time = time.time()

        # 先获取总数
        first_page = self.client.search_datasets(query=query, fq=fq, rows=1)
        total = first_page.get("count", 0)
        log.info("命中总数: %d", total)

        datasets: List[Dataset] = []
        for raw in self.client.iter_datasets(query=query, fq=fq, max_results=max_results):
            try:
                datasets.append(_parse_dataset(raw))
            except Exception as e:
                log.warning("解析数据集失败 id=%s: %s", raw.get("id"), e)

        elapsed = time.time() - start_time
        return SearchResult(
            query=f"{query} | fq={fq}",
            total_count=total,
            fetched_count=len(datasets),
            fetch_time_sec=round(elapsed, 2),
            datasets=datasets,
        )

    # ── 获取单个数据集详情 ───────────────────────────────────────────────────
    def get_dataset_detail(self, name_or_id: str) -> Optional[Dataset]:
        try:
            raw = self.client.get_dataset(name_or_id)
            return _parse_dataset(raw)
        except Exception as e:
            log.error("获取数据集详情失败 %s: %s", name_or_id, e)
            return None

    # ── 统计报告 ─────────────────────────────────────────────────────────────
    def generate_stats(self, query: str = "*:*") -> Dict:
        log.info("生成统计报告...")
        formats = self.client.facet_formats(query)
        orgs    = self.client.facet_organizations(query)
        return {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "query": query,
            "top_formats": formats,
            "top_organizations": orgs,
        }


# ─── 输出工具 ─────────────────────────────────────────────────────────────────
class OutputWriter:
    """将结构化数据写入 JSON / CSV / Excel。"""

    def __init__(self, output_dir: Path = OUTPUT_DIR):
        self.dir = output_dir

    def _path(self, filename: str) -> Path:
        return self.dir / filename

    # ── JSON ─────────────────────────────────────────────────────────────────
    def write_json(self, data: Any, filename: str) -> Path:
        path = self._path(filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        log.info("JSON 已写入: %s", path)
        return path

    # ── CSV ──────────────────────────────────────────────────────────────────
    def write_datasets_csv(self, datasets: List[Dataset], filename: str) -> Path:
        path = self._path(filename)
        fields = [
            "id", "name", "title", "organization", "org_type",
            "license_title", "num_resources", "num_tags",
            "metadata_created", "metadata_modified",
            "state", "tags", "groups", "url",
        ]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for ds in datasets:
                row = asdict(ds)
                row["tags"]   = "|".join(ds.tags)
                row["groups"] = "|".join(ds.groups)
                writer.writerow({k: row.get(k, "") for k in fields})
        log.info("CSV 已写入: %s  (%d 行)", path, len(datasets))
        return path

    def write_resources_csv(self, datasets: List[Dataset], filename: str) -> Path:
        path = self._path(filename)
        fields = [
            "dataset_id", "dataset_title", "resource_id", "resource_name",
            "format", "url", "size", "mimetype", "created", "last_modified",
        ]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for ds in datasets:
                for res in ds.resources:
                    writer.writerow({
                        "dataset_id":    ds.id,
                        "dataset_title": ds.title,
                        "resource_id":   res.id,
                        "resource_name": res.name,
                        "format":        res.format,
                        "url":           res.url,
                        "size":          res.size or "",
                        "mimetype":      res.mimetype or "",
                        "created":       res.created or "",
                        "last_modified": res.last_modified or "",
                    })
        log.info("Resources CSV 已写入: %s", path)
        return path

    # ── Excel（可选，需安装 openpyxl） ───────────────────────────────────────
    def write_excel(self, datasets: List[Dataset], filename: str) -> Optional[Path]:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            log.warning("openpyxl 未安装，跳过 Excel 输出。pip install openpyxl")
            return None

        wb = openpyxl.Workbook()

        # ── Sheet1: 数据集概览 ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Datasets"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        headers = [
            "ID", "Name", "Title", "Organization", "License",
            "# Resources", "# Tags", "Created", "Modified", "URL",
        ]
        ws1.append(headers)
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ds in datasets:
            ws1.append([
                ds.id, ds.name, ds.title,
                ds.organization or "", ds.license_title or "",
                ds.num_resources, ds.num_tags,
                ds.metadata_created or "", ds.metadata_modified or "",
                ds.url,
            ])

        # 自动列宽
        for col_idx, col in enumerate(ws1.columns, 1):
            max_len = max(len(str(c.value or "")) for c in col)
            ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        # ── Sheet2: 资源详情 ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Resources")
        res_headers = [
            "Dataset Title", "Resource Name", "Format",
            "Size (bytes)", "MIME Type", "URL", "Last Modified",
        ]
        ws2.append(res_headers)
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        for ds in datasets:
            for res in ds.resources:
                ws2.append([
                    ds.title, res.name, res.format,
                    res.size or "", res.mimetype or "",
                    res.url, res.last_modified or "",
                ])

        # ── Sheet3: 标签汇总 ──────────────────────────────────────────────
        ws3 = wb.create_sheet("Tag Frequency")
        tag_counter: Dict[str, int] = {}
        for ds in datasets:
            for tag in ds.tags:
                tag_counter[tag] = tag_counter.get(tag, 0) + 1
        ws3.append(["Tag", "Count"])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for tag, cnt in sorted(tag_counter.items(), key=lambda x: -x[1]):
            ws3.append([tag, cnt])

        path = self._path(filename)
        wb.save(path)
        log.info("Excel 已写入: %s", path)
        return path
